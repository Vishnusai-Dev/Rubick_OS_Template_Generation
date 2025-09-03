"""
Complete Streamlit app — single-file.
Saves/returns filled SKU template based on uploaded input, mapping & template.
"""

import streamlit as st
import pandas as pd
import openpyxl
import os
import re
from io import BytesIO
from typing import Optional, Tuple

# ───────────────────────── CONFIG / DEFAULT PATHS ─────────────────────────
DEFAULT_TEMPLATE_PATH = "sku-template (4).xlsx"
DEFAULT_MAPPING_PATH  = "Mapping - Automation.xlsx"

# Keys expected in mapping file (lowercased)
ATTR_KEY = "attributes"
TARGET_KEY = "fieldname"
MAND_KEY = "mandatoryornot"   # optional
TYPE_KEY = "fieldtype"       # optional
DUP_KEY  = "duplicatestobecreated"  # optional

# Predefined lists for size & color (lowercased)
SIZE_TOKENS = {
    "xxs","xs","s","m","l","xl","xxl","xxxl","2xl","3xl","4xl","free size","onesize",
    *[str(x) for x in range(20,61)]  # numeric sizes 20..60
}
COLOR_TOKENS = {
    "black","white","red","blue","green","yellow","pink","grey","gray","purple","orange","brown",
    "beige","maroon","navy","olive","khaki","gold","silver","cream","ivory","peach","cyan","teal",
    "turquoise","mustard","magenta","lavender","plum"
}

# image detection
IMAGE_EXT_RE = re.compile(r"(?i)\.(jpe?g|png|gif|bmp|webp|tiff?)$")
IMAGE_KEYWORDS = {"image","img","picture","photo","thumbnail","thumb","hero","front","back","url"}

# ─────────────────────────- HELPERS ─────────────────────────
def norm(s) -> str:
    if pd.isna(s):
        return ""
    return "".join(str(s).split()).lower()

def read_excelfile_from_uploaded(uploaded_file) -> Tuple[Optional[pd.ExcelFile], Optional[str], Optional[str]]:
    """
    Read uploaded file into pandas.ExcelFile with the correct engine.
    Returns (ExcelFile or None, detected_ext, error_message or None)
    """
    if uploaded_file is None:
        return None, None, "No file provided"

    fname = uploaded_file.name
    ext = os.path.splitext(fname)[1].lower()
    data = uploaded_file.read()
    buf = BytesIO(data)

    try:
        if ext == ".xls":
            # NOTE: requires xlrd==1.2.0 installed
            xl = pd.ExcelFile(buf, engine="xlrd")
        else:
            xl = pd.ExcelFile(buf, engine="openpyxl")
        return xl, ext, None
    except Exception as e:
        return None, ext, f"Failed to read Excel file {fname}: {e}"

def _select_sheet_and_rows(xl: pd.ExcelFile, template_type: str) -> Tuple[Optional[str], Optional[int], Optional[int]]:
    """
    Given pandas.ExcelFile and template_type, return (sheet_name, header_row_1based, data_start_row_1based).
    Returns (None, None, None) if sheet not found / insufficient sheets.
    """
    if template_type == "Amazon":
        if "Template" not in xl.sheet_names:
            return None, None, None
        return "Template", 2, 4
    if template_type == "Flipkart":
        if len(xl.sheet_names) < 3:
            return None, None, None
        return xl.sheet_names[2], 1, 5
    if template_type == "Myntra":
        if len(xl.sheet_names) < 2:
            return None, None, None
        return xl.sheet_names[1], 3, 4
    if template_type == "Ajio":
        if len(xl.sheet_names) < 3:
            return None, None, None
        return xl.sheet_names[2], 2, 3
    if template_type == "TataCliq":
        if len(xl.sheet_names) < 1:
            return None, None, None
        return xl.sheet_names[0], 4, 6
    # General default: first sheet, header row 1, data starts row 2
    return xl.sheet_names[0], 1, 2

def rebase_header(df_raw: pd.DataFrame, header_row_1b: int, data_start_row_1b: int) -> pd.DataFrame:
    """Make header_row become columns, and data start at data_start_row_1b. Return cleaned df."""
    hdr_idx = max(header_row_1b - 1, 0)
    data_idx = max(data_start_row_1b - 1, 0)
    if hdr_idx >= len(df_raw):
        raise ValueError("Header row index exceeds sheet length")
    headers = df_raw.iloc[hdr_idx].astype(str).tolist()
    # dedupe headers
    seen = {}
    dedup = []
    for h in headers:
        h0 = "" if (pd.isna(h) or str(h).strip().lower()=="nan") else str(h).strip()
        if h0 in seen:
            seen[h0] += 1
            dedup.append(f"{h0}_{seen[h0]}")
        else:
            seen[h0] = 1
            dedup.append(h0)
    df = df_raw.iloc[data_idx:].copy()
    df.columns = dedup
    df = df.dropna(how="all")
    df = df.loc[:, ~df.columns.to_series().apply(lambda c: df[c].isna().all())]
    df.reset_index(drop=True, inplace=True)
    return df

def _ratio_match(series: pd.Series, token_set:set, extra_check=None) -> float:
    """Return fraction of non-empty values that match token_set or extra_check."""
    if series is None or series.empty:
        return 0.0
    s = series.dropna().astype(str).str.strip()
    if s.empty:
        return 0.0
    def hit(x):
        up = x.upper()
        if up in (t.upper() for t in token_set):
            return True
        if extra_check:
            return extra_check(x)
        return False
    return s.apply(hit).mean()

def looks_like_numeric_size(v: str) -> bool:
    # 1-3 digit numeric sizes (e.g., 28, 100)
    return bool(re.fullmatch(r"\d{1,3}", v.strip()))

def looks_like_color_value(v: str) -> bool:
    # if any word token is in COLOR_TOKENS
    v2 = re.sub(r"[^\w ]", " ", v).strip()
    parts = [p for p in v2.split() if p]
    return any(p.lower() in COLOR_TOKENS for p in parts)

def find_size_column(df: pd.DataFrame) -> Tuple[Optional[str], float]:
    # header-first
    for c in df.columns:
        if "size" in norm(c):
            return c, 1.0
    # value-led
    best_col, best_ratio = None, 0.0
    for c in df.columns:
        r = _ratio_match(df[c], SIZE_TOKENS, looks_like_numeric_size)
        if r > best_ratio:
            best_col, best_ratio = c, r
    return (best_col, best_ratio) if best_ratio >= 0.35 else (None, 0.0)

def find_color_column(df: pd.DataFrame) -> Tuple[Optional[str], float]:
    for c in df.columns:
        n = norm(c)
        if "color" in n or "colour" in n:
            return c, 1.0
    best_col, best_ratio = None, 0.0
    for c in df.columns:
        r = _ratio_match(df[c], COLOR_TOKENS, looks_like_color_value)
        if r > best_ratio:
            best_col, best_ratio = c, r
    return (best_col, best_ratio) if best_ratio >= 0.35 else (None, 0.0)

def is_image_column(col_header_norm: str, series: pd.Series) -> bool:
    header_hit = any(k in col_header_norm for k in IMAGE_KEYWORDS)
    sample = series.dropna().astype(str).head(20)
    ratio  = sample.str.contains(IMAGE_EXT_RE).mean() if not sample.empty else 0.0
    return header_hit or ratio >= 0.30

# ───────────────────────── Mapping loader (optional upload) ─────────────────────────
def load_mapping_from_uploaded(uploaded_mapping_file) -> Tuple[pd.DataFrame, list]:
    """
    Returns (mapping_df (with normalized column names), client_names_list)
    If uploaded_mapping_file is None and default mapping path exists, try loading that.
    """
    if uploaded_mapping_file:
        buf = BytesIO(uploaded_mapping_file.read())
        try:
            xl = pd.ExcelFile(buf, engine="openpyxl")
        except Exception as e:
            st.error(f"Failed to read uploaded mapping file: {e}")
            return pd.DataFrame(), []
    else:
        if not os.path.exists(DEFAULT_MAPPING_PATH):
            return pd.DataFrame(), []
        try:
            xl = pd.ExcelFile(DEFAULT_MAPPING_PATH, engine="openpyxl")
        except Exception as e:
            st.error(f"Failed to read default mapping file {DEFAULT_MAPPING_PATH}: {e}")
            return pd.DataFrame(), []

    # find mapping sheet
    map_sheet = next((s for s in xl.sheet_names if "mapping" in norm(s)), xl.sheet_names[0])
    mapping_df = xl.parse(map_sheet)
    # lowercase & normalize column names
    mapping_df.rename(columns={c: norm(c) for c in mapping_df.columns}, inplace=True)
    if ATTR_KEY not in mapping_df.columns or TARGET_KEY not in mapping_df.columns:
        # keep mapping_df but warn
        st.warning(f"Mapping sheet found ({map_sheet}) but doesn't include required columns '{ATTR_KEY}' and '{TARGET_KEY}'. Mapping will be limited.")
    mapping_df["__attr_key"] = mapping_df[ATTR_KEY].apply(norm) if ATTR_KEY in mapping_df.columns else mapping_df.iloc[:,0].astype(str).apply(norm)
    # try to read client sheet names if present
    client_sheet = next((s for s in xl.sheet_names if "mappedclientname" in norm(s)), None)
    client_names = []
    if client_sheet:
        raw = xl.parse(client_sheet, header=None)
        client_names = [str(x).strip() for x in raw.values.flatten() if pd.notna(x) and str(x).strip()]
    return mapping_df, client_names

# ───────────────────────── Process file → fill template → return BytesIO ────────────────────────
def process_and_build_output(
    uploaded_input_file,
    template_type: str,
    mode: str,
    uploaded_mapping_file=None,
    uploaded_template_file=None
) -> Tuple[Optional[BytesIO], list]:
    """
    Main processing function. Returns (BytesIO_of_output, log_lines)
    """
    logs = []
    # read input Excel
    xl_input, ext, err = read_excelfile_from_uploaded(uploaded_input_file)
    if err:
        logs.append(err)
        return None, logs

    sheet_name, hdr_1b, data_1b = _select_sheet_and_rows(xl_input, template_type)
    if sheet_name is None:
        logs.append("Selected template type requires a sheet that is missing in the uploaded input.")
        return None, logs
    logs.append(f"Using sheet '{sheet_name}' (header row: {hdr_1b}, data starts: {data_1b})")

    # read sheet raw, then rebase header
    try:
        df_raw = xl_input.parse(sheet_name, header=None)
    except Exception as e:
        logs.append(f"Failed to parse sheet {sheet_name}: {e}")
        return None, logs

    try:
        src_df = rebase_header(df_raw, hdr_1b, data_1b)
    except Exception as e:
        logs.append(f"Failed to rebase header: {e}")
        return None, logs
    logs.append(f"Input data shape after rebasing: {src_df.shape}")

    # load mapping
    mapping_df, client_names = load_mapping_from_uploaded(uploaded_mapping_file)
    if not mapping_df.empty:
        logs.append(f"Loaded mapping sheet with {len(mapping_df)} rows; clients found: {', '.join(client_names) if client_names else 'none'}")
    else:
        logs.append("No mapping loaded (will use auto-mapping behavior)")

    # build columns_meta (mapping vs auto)
    columns_meta = []
    if mode == "Mapping" and not mapping_df.empty:
        for col in src_df.columns:
            col_key = norm(col)
            matches = mapping_df[mapping_df["__attr_key"] == col_key] if "__attr_key" in mapping_df.columns else pd.DataFrame()
            if not matches.empty:
                row3 = matches.iloc[0][MAND_KEY] if MAND_KEY in matches.columns else "mandatory"
                row4 = matches.iloc[0][TYPE_KEY] if TYPE_KEY in matches.columns else "string"
            else:
                row3 = row4 = "Not Found"
            columns_meta.append({"src": col, "out": col, "row3": row3, "row4": row4})
            for _, row in matches.iterrows():
                if DUP_KEY in mapping_df.columns and str(row.get(DUP_KEY,"")).strip().lower().startswith("yes"):
                    new_header = row.get(TARGET_KEY, col) if TARGET_KEY in mapping_df.columns else col
                    if pd.notna(new_header) and new_header != col:
                        columns_meta.append({
                            "src": col,
                            "out": new_header,
                            "row3": row.get(MAND_KEY, row3) if MAND_KEY in mapping_df.columns else row3,
                            "row4": row.get(TYPE_KEY, row4) if TYPE_KEY in mapping_df.columns else row4
                        })
    else:
        # Auto-mapping
        for col in src_df.columns:
            dtype = "imageurlarray" if is_image_column(norm(col), src_df[col]) else "string"
            columns_meta.append({"src": col, "out": col, "row3": "mandatory", "row4": dtype})

    logs.append(f"Columns prepared for writing: {len(columns_meta)}")

    # identify size & color columns
    size_col, size_ratio = find_size_column(src_df)
    color_col, color_ratio = find_color_column(src_df)
    logs.append(f"Size column detected: {size_col} (ratio {size_ratio:.2f})")
    logs.append(f"Color column detected: {color_col} (ratio {color_ratio:.2f})")

    # copy option columns (full columns) — ensure length matches src_df
    option1_series = src_df[size_col] if size_col and size_col in src_df.columns else pd.Series([""]*len(src_df))
    option2_series = src_df[color_col] if color_col and color_col in src_df.columns else pd.Series([""]*len(src_df))

    # load template workbook (uploaded or default)
    try:
        if uploaded_template_file:
            temp_bytes = uploaded_template_file.read()
            wb = openpyxl.load_workbook(BytesIO(temp_bytes))
        else:
            if not os.path.exists(DEFAULT_TEMPLATE_PATH):
                logs.append(f"Default template not found at {DEFAULT_TEMPLATE_PATH}")
                return None, logs
            wb = openpyxl.load_workbook(DEFAULT_TEMPLATE_PATH)
    except Exception as e:
        logs.append(f"Failed to load template workbook: {e}")
        return None, logs

    # ensure Values & Types sheets exist (try common names)
    values_name = next((s for s in wb.sheetnames if s.lower().startswith("value")), None)
    types_name  = next((s for s in wb.sheetnames if s.lower().startswith("type")), None)
    log_name    = "Log"
    if not values_name or not types_name:
        logs.append("Template must contain sheets starting with 'Values' and 'Types' (case-insensitive).")
        return None, logs
    ws_vals = wb[values_name]
    ws_types = wb[types_name]

    # clear those sheets
    for ws in (ws_vals, ws_types):
        max_r, max_c = ws.max_row, ws.max_column
        if max_r >= 1 and max_c >= 1:
            for r in range(1, max_r+1):
                for c in range(1, max_c+1):
                    ws.cell(row=r, column=c).value = None

    # Write main mapped/auto-mapped columns to Values and Types
    for j, meta in enumerate(columns_meta, start=1):
        header_display = str(meta["out"])
        ws_vals.cell(row=1, column=j, value=header_display)
        col_name = meta["src"]
        col_series = src_df[col_name] if col_name in src_df.columns else pd.Series([None]*len(src_df))
        for i, v in enumerate(col_series.tolist(), start=2):
            cell = ws_vals.cell(row=i, column=j)
            if pd.isna(v) or (isinstance(v, str) and v.strip()==""):
                cell.value = None
            else:
                if str(meta["row4"]).lower() in ("string", "imageurlarray"):
                    cell.value = str(v)
                    cell.number_format = "@"
                else:
                    cell.value = v
        # Types sheet (shifted by +2)
        tcol = j + 2
        ws_types.cell(row=1, column=tcol, value=header_display)
        ws_types.cell(row=2, column=tcol, value=header_display)
        ws_types.cell(row=3, column=tcol, value=meta.get("row3",""))
        ws_types.cell(row=4, column=tcol, value=meta.get("row4",""))

    # Append Option 1 & Option 2 to Values
    opt1_col = len(columns_meta) + 1
    opt2_col = len(columns_meta) + 2
    ws_vals.cell(row=1, column=opt1_col, value="Option 1")
    ws_vals.cell(row=1, column=opt2_col, value="Option 2")
    for i, v in enumerate(option1_series.tolist(), start=2):
        ws_vals.cell(row=i, column=opt1_col, value=(None if pd.isna(v) or str(v).strip()=="" else v))
    for i, v in enumerate(option2_series.tolist(), start=2):
        ws_vals.cell(row=i, column=opt2_col, value=(None if pd.isna(v) or str(v).strip()=="" else v))

    # Append Option 1 & 2 to Types
    t1_col = opt1_col + 2
    t2_col = opt2_col + 2
    # Option 1 Types
    ws_types.cell(row=1, column=t1_col, value="Option 1")
    ws_types.cell(row=2, column=t1_col, value="Option 1")
    ws_types.cell(row=3, column=t1_col, value="non mandatory")
    ws_types.cell(row=4, column=t1_col, value="string")
    unique_opt1 = pd.Series([str(x) for x in option1_series.dropna().astype(str).tolist() if str(x).strip()])
    unique_opt2 = pd.Series([str(x) for x in option2_series.dropna().astype(str).tolist() if str(x).strip()])
    for idx_val, v in enumerate(sorted(unique_opt1.unique().tolist()), start=5):
        ws_types.cell(row=idx_val, column=t1_col, value=v)
    # Option 2 Types
    ws_types.cell(row=1, column=t2_col, value="Option 2")
    ws_types.cell(row=2, column=t2_col, value="Option 2")
    ws_types.cell(row=3, column=t2_col, value="non mandatory")
    ws_types.cell(row=4, column=t2_col, value="string")
    for idx_val, v in enumerate(sorted(unique_opt2.unique().tolist()), start=5):
        ws_types.cell(row=idx_val, column=t2_col, value=v)

    # Add a small Log sheet
    if log_name in wb.sheetnames:
        log_ws = wb[log_name]
        for r in range(1, log_ws.max_row+1):
            for c in range(1, log_ws.max_column+1):
                log_ws.cell(row=r, column=c).value = None
    else:
        log_ws = wb.create_sheet(log_name)
    # write logs
    for i, line in enumerate(["Processing log:"] + logs, start=1):
        log_ws.cell(row=i, column=1, value=line)
    log_ws.cell(row=len(logs)+3, column=1, value=f"Detected size column: {size_col} (ratio {size_ratio:.2f})")
    log_ws.cell(row=len(logs)+4, column=1, value=f"Detected color column: {color_col} (ratio {color_ratio:.2f})")

    # Save to BytesIO and return
    out_buf = BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    logs.insert(0, f"Output generated with template '{values_name}'/'{types_name}'.")
    return out_buf, logs

# ───────────────────────── STREAMLIT UI ─────────────────────────
st.set_page_config(page_title="SKU Template Automation (complete)", layout="wide")
st.title("SKU Template Automation — Complete Tool")

st.markdown(
    "Upload input file (xls/xlsx/xlsm). Optionally upload a Mapping workbook and/or Template workbook. "
    "If you don't upload mapping or template, local files named "
    f"`{DEFAULT_MAPPING_PATH}` and `{DEFAULT_TEMPLATE_PATH}` will be used if present."
)

# Inputs: template & mapping upload optional
uploaded_template_file = st.file_uploader("Upload SKU template (optional) — must contain Values & Types sheets", type=["xlsx","xlsm"])
uploaded_mapping_file = st.file_uploader("Upload Mapping workbook (optional)", type=["xlsx","xlsm","xls"])

# Mode & template type
mode = st.selectbox("Mode", ["Auto-Mapping", "Mapping"])
template_type = st.selectbox("Type of input (marketplace)", ["General","Amazon","Flipkart","Myntra","Ajio","TataCliq"])

# Input file uploader (required)
uploaded_input_file = st.file_uploader("Upload Input Excel (xls/xlsx/xlsm) — required", type=["xls","xlsx","xlsm"])

if uploaded_input_file:
    with st.spinner("Processing…"):
        out_buf, logs = process_and_build_output(
            uploaded_input_file,
            template_type=template_type,
            mode=("Mapping" if mode=="Mapping" else "Auto-Mapping"),
            uploaded_mapping_file=uploaded_mapping_file,
            uploaded_template_file=uploaded_template_file
        )
    # show logs
    if logs:
        st.text_area("Process log", "\n".join(logs), height=200)

    if out_buf:
        st.success("✅ Output generated — download below.")
        st.download_button(
            "📥 Download Filled Template",
            data=out_buf,
            file_name="output_template_filled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("Processing failed — see log above.")

st.markdown("---")
st.caption("Built for Rubick.ai | By Vishnu Sai")
